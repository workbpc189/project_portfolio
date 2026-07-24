import openpyxl, os
import pyinputplus as pyip
from pathlib import Path
import re

# Function that finds the file path for each drawing.
def find_file_path(folder_name):
    for folders, subfolders, files in os.walk(r"C:\Users"):
      for subfolder in subfolders:
        if subfolder == folder_name:
          full_path = os.path.join(folders, folder_name)
    return full_path


# Function to create Excel files.
def create_excel_file(foldername):
  wb = openpyxl.Workbook()
  sheet = wb.active
  sheet.title = "Drawings"
  sheet["A1"] = "Drawing Number"
  sheet["B1"] = "Drawing Name"
  sheet["C1"] = "Drawing Date"
  sheet["D1"] = "File Hyperlink"
  wb.save("drawings.xlsx")
  for folders, subfolders, files in os.walk(foldername):
    edit_excel_file(files, "drawings.xlsx")


# Function that edits Excel files.
def edit_excel_file(file_list, xl_name):
  wb = openpyxl.load_workbook(xl_name)
  sheet = wb.active
  for file in file_list:
    number = file_list.index(file) + 1
    name = pyip.inputStr(f"Enter a name for the drawing.\n")
    date = pyip.inputStr(f"Enter the date for when the drawing was made.\n")
    n = number + 1
    sheet[f"A{n}"] = number
    sheet[f"B{n}"] = name
    sheet[f"C{n}"] = date
    make_hyperlink(pyip.inputStr(f"Paste the hyperlink for the drawing.\n"), n, xl_name)
  wb.save("drawings.xlsx")


# Function to re-format string into hyperlink.
def make_hyperlink(link, cell_num, excel_name):
  wb = openpyxl.load_workbook(excel_name)
  sheet = wb.active
  cell = sheet[f"D{cell_num}"]
  cell.value = "Click Here"
  cell.hyperlink = link
  wb.save(excel_name)


# Main Program Loop
x = pyip.inputStr(f"Enter the name of the folder with your drawings in it.\n")
file_path = find_file_path(x)
print(file_path)
create_excel_file(file_path)
print("Your excel file has been created.")


"""
NOTES:
Need to make the finishing touches so it's fully user interactive and not partially hard-coded.
Would also be a good idea to build some sort of UI so non-developers can use it.
"""